#!/usr/bin/env python3
"""
Agent1: 基于客户模板生成三张表（01/02/03）- 一级费用科目使用API现有名称，其他用新名称
"""

from __future__ import annotations

import argparse
import json
import random
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import requests
import websocket

BASE_URL = "https://cst.uf-tree.com"

# 支持的浏览器 CDP 端口
BROWSERS = [
    {"name": "Edge", "port": 9223, "url": "http://localhost:9223/json"},
    {"name": "Chrome", "port": 18800, "url": "http://localhost:18800/json"},
]

DOC_TYPES = ["报销单", "借款单", "批量付款单", "申请单"]

# 二级、三级费用科目名称词库（新的唯一名称）
SECONDARY_SUBJECT_NAMES = [
    "人员费用", "设备耗材", "差旅交通", "会议培训", "外包服务",
    "软件许可", "云资源", "办公场地", "通讯费用", "招待费用"
]

THIRD_SUBJECT_NAMES = [
    "机票", "火车票", "酒店住宿", "餐饮补贴", "市内交通",
    "打印复印", "办公用品", "设备维护", "网络费用", "电话费",
    "快递费", "资料费", "咨询费", "审计费", "培训费"
]

# 单据名称词库（全新的，与系统现有不同）
DOC_NAME_TEMPLATES = [
    "项目{name}报销单", "部门{name}申请单", "{name}费用报销单", 
    "日常{name}报销单", "专项{name}付款单", "临时{name}借款单",
    "行政{name}申请单", "业务{name}报销单"
]

DOC_TYPE_CHOICES = [
    ("报销单", True),    # (单据类型, 是否有人员)
    ("报销单", False),
    ("批量付款单", True),
    ("借款单", False),
    ("申请单", False),
]

@dataclass
class Auth:
    token: str
    company_id: int
    company_name: str


def find_browser():
    """自动检测可用的浏览器"""
    for browser in BROWSERS:
        try:
            pages = requests.get(browser["url"], timeout=6).json()
            has_cst = any("cst.uf-tree.com" in p.get("url", "") for p in pages)
            if has_cst:
                return browser
        except Exception:
            continue
    return None


def get_auth_from_edge(token_override=None, company_id_override=None, company_name_override=None) -> Auth:
    if token_override and company_id_override:
        return Auth(
            token=token_override,
            company_id=int(company_id_override),
            company_name=company_name_override or f"company_{company_id_override}",
        )

    browser = find_browser()
    if not browser:
        raise RuntimeError("未检测到可用浏览器，请先登录 https://cst.uf-tree.com")

    pages = requests.get(browser["url"], timeout=6).json()
    ws_url = None
    for p in pages:
        if "cst.uf-tree.com" in p.get("url", ""):
            ws_url = p.get("webSocketDebuggerUrl")
            break
    if not ws_url:
        raise RuntimeError("未找到财税通页面")

    ws = websocket.create_connection(ws_url, timeout=10, suppress_origin=True)
    ws.send(json.dumps({
        "id": 1,
        "method": "Runtime.evaluate",
        "params": {
            "expression": "(function(){const v=localStorage.getItem('vuex');if(!v)return null;const d=JSON.parse(v);return {token:d.user?.token, companyId:d.user?.company?.id, companyName:d.user?.company?.name};})()",
            "returnByValue": True,
        },
    }))
    raw = ws.recv()
    ws.close()

    value = json.loads(raw).get("result", {}).get("result", {}).get("value")
    if not value or not value.get("token") or not value.get("companyId"):
        raise RuntimeError("读取 token/companyId 失败（请确认已登录）")

    return Auth(
        token=value["token"],
        company_id=value["companyId"],
        company_name=value.get("companyName", ""),
    )


def api_get(auth: Auth, endpoint: str, params: dict):
    return requests.get(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        params=params,
        timeout=15,
    ).json()


def api_post(auth: Auth, endpoint: str, payload: dict):
    return requests.post(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        json=payload,
        timeout=15,
    ).json()


def unique_names(seq: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in seq:
        x = (x or "").strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def fetch_sources(auth: Auth) -> Dict:
    """从系统获取基础数据 - 包括一级费用科目"""
    deps_resp = api_get(auth, "/api/member/department/queryDepartments", {"companyId": auth.company_id})
    deps = deps_resp.get("result", []) or []
    dep_names = unique_names([d.get("title") for d in deps])

    role_resp = api_get(auth, "/api/member/role/get/tree", {"companyId": auth.company_id})
    role_tree = role_resp.get("result", []) or []
    role_names = []
    for cat in role_tree:
        for c in (cat.get("children") or []):
            if c.get("name"):
                role_names.append(c.get("name"))
    role_names = unique_names(role_names)

    wf_resp = api_get(auth, "/api/bpm/workflow/queryWorkFlow", {
        "companyId": auth.company_id,
        "t": int(time.time() * 1000),
    })
    workflows = wf_resp.get("result", []) or []
    wf = next((x for x in workflows if x.get("tpName") == "通用审批流"), None)

    # 获取一级费用科目（从API）
    fee_resp = api_get(auth, "/api/bill/feeTemplate/queryFeeTemplate", {
        "companyId": auth.company_id,
        "status": 1,
        "pageSize": 1000,
    })
    fee_tree = fee_resp.get("result", []) or []
    # 只取一级科目（parentId == -1）
    primary_subjects = [p.get("name") for p in fee_tree if p.get("parentId") == -1 and p.get("name")]
    
    return {
        "departments": dep_names,
        "roles": role_names,
        "primary_subjects": primary_subjects,  # API中现有的一级科目
        "workflow_name": (wf or {}).get("tpName", "通用审批流"),
    }


def generate_sheet1(company_name: str, departments: List[str], employee_count: int = 10):
    """生成员工表 - 全新的随机员工姓名"""
    surnames = ["张", "李", "王", "赵", "刘", "陈", "周", "吴", "林", "黄", "徐", "孙", "马", "朱", "胡"]
    given = ["晨", "悦", "宁", "航", "嘉", "然", "宇", "琳", "涛", "雪", "瑞", "欣", "怡", "泽", "睿"]

    names_used, phones_used = set(), set()
    rows = []

    for i in range(employee_count):
        while True:
            name = f"{random.choice(surnames)}{random.choice(given)}{i+1}"
            if name not in names_used:
                names_used.add(name)
                break

        while True:
            phone = f"1{random.randint(30,99)}{random.randint(1000,9999)}{random.randint(1000,9999)}"
            if phone not in phones_used:
                phones_used.add(phone)
                break

        dep = random.choice(departments) if departments else company_name
        rows.append({
            "姓名": name,
            "手机号": phone,
            "企业名称": company_name,
            "一级部门名称": dep,
            "二级部门": "",
            "备注": "唯一性已校验",
        })
    return rows


def generate_sheet2(users: List[str], primary_subjects: List[str]):
    """
    生成费用科目配置表
    - 一级费用科目：使用API中现有的名称
    - 二级费用科目：全新的唯一名称
    - 三级费用科目：全新的唯一名称
    - 归属单据名称：全新的唯一名称
    """
    # 如果没有从API获取到一级科目，使用默认值
    if not primary_subjects:
        primary_subjects = ["运营增长", "AI工程", "客户交付", "财税支持"]
    
    # 随机选择5个单据模板（每个对应不同的单据类型）
    selected_docs = []
    used_names = set()
    
    for i in range(5):
        doc_type, has_people = DOC_TYPE_CHOICES[i % len(DOC_TYPE_CHOICES)]
        # 生成唯一的单据名称
        while True:
            name_keyword = random.choice(["差旅", "日常", "采购", "项目", "行政", "业务", "专项", "临时"])
            template = random.choice(DOC_NAME_TEMPLATES)
            doc_name = template.format(name=name_keyword)
            if doc_name not in used_names:
                used_names.add(doc_name)
                break
        selected_docs.append({
            "type": doc_type,
            "name": doc_name,
            "has_people": has_people
        })

    # 为每个单据创建费用科目结构
    rows = []
    used_secondary = set()
    used_third = set()
    
    for doc in selected_docs:
        # 从API现有的一级科目中选择一个
        primary = random.choice(primary_subjects)
        
        # 为每个单据创建2个二级科目
        for _ in range(2):
            # 选择二级科目（新的唯一名称）
            available_secondary = [s for s in SECONDARY_SUBJECT_NAMES if s not in used_secondary]
            if not available_secondary:
                used_secondary.clear()
                available_secondary = SECONDARY_SUBJECT_NAMES
            secondary = random.choice(available_secondary)
            used_secondary.add(secondary)
            
            # 为每个二级创建2-3个三级科目（新的唯一名称）
            third_count = random.randint(2, 3)
            for _ in range(third_count):
                available_third = [t for t in THIRD_SUBJECT_NAMES if t not in used_third]
                if not available_third:
                    used_third.clear()
                    available_third = THIRD_SUBJECT_NAMES
                third = random.choice(available_third)
                used_third.add(third)
                
                # 分配人员（如果单据需要人员）
                people = ""
                if doc["has_people"] and len(users) >= 2:
                    people = "，".join(random.sample(users, 2))
                
                rows.append({
                    "一级费用科目": primary,  # 使用 API 中现有的名称
                    "二级费用科目": f"{secondary}_{random.randint(1000,9999)}",  # 新的唯一名称
                    "三级费用科目": f"{third}_{random.randint(1000,9999)}",  # 新的唯一名称
                    "四级费用科目": "",
                    "归属单据类型": doc["type"],
                    "归属单据名称": doc["name"],  # 新的唯一名称
                    "单据适配人员": people,
                    "是否执行": "是",
                    "备注": f"单据:{doc['name']}",
                })
    
    # 按层级排序
    rows.sort(key=lambda r: (r["一级费用科目"], r["二级费用科目"], r["三级费用科目"]))
    return rows


def build_sheet3_from_sheet2(sheet2_rows: List[Dict], roles: List[str], users: List[str], deps: List[str], workflow_name: str, inherit_group_visual=True):
    """从 Sheet2 生成 Sheet3 单据表"""
    # 汇总同名单据
    agg: Dict[str, Dict] = {}

    for r in sheet2_rows:
        name = (r.get("归属单据名称") or "").strip()
        if not name:
            continue

        doc_type = (r.get("归属单据类型") or "").strip()
        has_people = bool((r.get("单据适配人员") or "").strip())
        if name not in agg:
            agg[name] = {"doc_type": doc_type, "any_people": False}
        agg[name]["any_people"] = agg[name]["any_people"] or has_people

    group_map = {
        "报销单": "报销类单据",
        "批量付款单": "付款类单据",
        "借款单": "借款类单据",
        "申请单": "申请类单据",
    }

    out = []
    for i, (doc_name, info) in enumerate(agg.items()):
        dt = info["doc_type"]
        if dt not in DOC_TYPES:
            continue

        # 可见范围：全员（避免导入失败）
        vis_type = "全员"
        vis_obj = ""

        # 费用限制规则
        if dt in ["报销单", "批量付款单"] and info["any_people"]:
            limit_mode = "费用角色限制"
            target = f"费用角色:{doc_name}"
        else:
            limit_mode = ""
            target = ""
        
        out.append({
            "单据分组（一级目录）": group_map[dt],
            "单据大类（二级目录）": dt,
            "单据模板名称": doc_name,  # 保持原名称，不加时间戳
            "可见范围类型": vis_type,
            "可见范围对象": vis_obj,
            "备注": "全新生成的单据模板",
            "费用限制方式": limit_mode,
            "对应末级科目/对应费用角色": target,
            "审批流": workflow_name,
        })

    # 视觉继承
    if inherit_group_visual and out:
        last = None
        for r in out:
            g = r["单据分组（一级目录）"]
            if g == last:
                r["单据分组（一级目录）"] = ""
            else:
                last = g

    return out


def clear_range(ws, row_start=4, row_end=220, col_start=1, col_end=8):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(r, c).value = None


def unmerge_all(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))


def clear_data_validations(ws):
    ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()


def merge_same(ws, col: int, start_row: int, end_row: int):
    r = start_row
    while r <= end_row:
        v = ws.cell(r, col).value
        e = r
        while e + 1 <= end_row and ws.cell(e + 1, col).value == v:
            e += 1
        if e > r and v not in (None, ""):
            ws.merge_cells(start_row=r, start_column=col, end_row=e, end_column=col)
        r = e + 1


def write_by_customer_template(template_path: Path, out_path: Path, auth: Auth, sources: Dict, inherit_group_visual=True):
    wb = openpyxl.load_workbook(template_path)
    ws1 = wb["01_添加员工"]
    ws2 = wb["02_费用科目配置"]
    ws3 = wb["03_单据表"]
    wsb = wb["基础数据"]

    for ws in [ws1, ws2, ws3]:
        unmerge_all(ws)
        clear_data_validations(ws)

    clear_range(ws1, 4, 220, 1, 8)
    clear_range(ws2, 4, 220, 1, 9)
    clear_range(ws3, 4, 220, 1, 8)

    # 生成数据
    sheet1_rows = generate_sheet1(auth.company_name, sources["departments"], employee_count=10)
    sheet1_names = [r["姓名"] for r in sheet1_rows]
    sheet2_rows = generate_sheet2(sheet1_names, sources["primary_subjects"])  # 传递API中现有的一级科目
    sheet3_rows = build_sheet3_from_sheet2(
        sheet2_rows,
        roles=sources["roles"],
        users=sheet1_names,
        deps=sources["departments"],
        workflow_name=sources["workflow_name"],
        inherit_group_visual=inherit_group_visual,
    )

    # Write Sheet1
    r = 4
    for i, row in enumerate(sheet1_rows, start=1):
        ws1.cell(r, 1).value = i
        ws1.cell(r, 2).value = "是"
        ws1.cell(r, 3).value = row["姓名"]
        ws1.cell(r, 4).value = row["手机号"]
        ws1.cell(r, 5).value = row["企业名称"]
        ws1.cell(r, 6).value = row["一级部门名称"]
        ws1.cell(r, 7).value = row["二级部门"]
        ws1.cell(r, 8).value = row["备注"]
        r += 1

    # Write Sheet2
    r = 4
    for i, row in enumerate(sheet2_rows, start=1):
        ws2.cell(r, 1).value = i
        ws2.cell(r, 2).value = row["是否执行"]
        ws2.cell(r, 3).value = row["一级费用科目"]
        ws2.cell(r, 4).value = row["二级费用科目"]
        ws2.cell(r, 5).value = row["三级费用科目"]
        ws2.cell(r, 6).value = row["归属单据名称"]
        ws2.cell(r, 7).value = row["单据适配人员"]
        ws2.cell(r, 8).value = row["四级费用科目"]
        ws2.cell(r, 9).value = row["备注"]
        r += 1
    end2 = r - 1

    merge_same(ws2, 3, 4, end2)
    merge_same(ws2, 4, 4, end2)

    # Write Sheet3
    r = 4
    for i, row in enumerate(sheet3_rows, start=1):
        ws3.cell(r, 1).value = i
        ws3.cell(r, 2).value = "是"
        ws3.cell(r, 3).value = row["单据分组（一级目录）"]
        ws3.cell(r, 4).value = row["单据大类（二级目录）"]
        ws3.cell(r, 5).value = row["单据模板名称"]
        ws3.cell(r, 6).value = row["可见范围类型"]
        ws3.cell(r, 7).value = row["可见范围对象"]
        ws3.cell(r, 8).value = row["备注"]
        r += 1

    # 基础数据
    clear_range(wsb, 2, 30, 1, 7)
    wsb.cell(1, 1).value = "可见范围类型"
    wsb.cell(1, 3).value = "角色"
    wsb.cell(1, 4).value = "员工"
    wsb.cell(1, 5).value = "部门"
    wsb.cell(1, 6).value = "单据大类"
    wsb.cell(1, 7).value = "是否创建"
    for i in range(10):
        rr = 2 + i
        wsb.cell(rr, 1).value = "样本"
        wsb.cell(rr, 3).value = sources["roles"][i % len(sources["roles"])] if sources["roles"] else ""
        wsb.cell(rr, 4).value = sheet1_names[i % len(sheet1_names)] if sheet1_names else ""
        wsb.cell(rr, 5).value = sources["departments"][i % len(sources["departments"])] if sources["departments"] else ""
        wsb.cell(rr, 6).value = DOC_TYPES[i % len(DOC_TYPES)]
        wsb.cell(rr, 7).value = "是"

    for ws in [ws1, ws2, ws3, wsb]:
        clear_data_validations(ws)
        ws.auto_filter.ref = None

    ws1.auto_filter.ref = "B3"
    ws2.auto_filter.ref = "B3"
    ws3.auto_filter.ref = "B3"

    wb.save(out_path)

    # 输出报告
    report = {
        "companyId": auth.company_id,
        "companyName": auth.company_name,
        "output": str(out_path),
        "counts": {
            "departments": len(sources["departments"]),
            "roles": len(sources["roles"]),
            "primary_subjects": len(sources["primary_subjects"]),
            "sheet1_rows": len(sheet1_rows),
            "sheet2_rows": len(sheet2_rows),
            "sheet3_rows": len(sheet3_rows),
        },
        "generated_names": {
            "单据名称": list(set(r["归属单据名称"] for r in sheet2_rows)),
            "一级科目(API现有)": sources["primary_subjects"],
        }
    }
    report_path = out_path.with_suffix(".report.json")
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report_path


def main():
    parser = argparse.ArgumentParser(description="基于客户模板生成三表（Agent1）- 一级科目用API现有名称")
    parser.add_argument("--template", required=True, help="客户模板xlsx路径")
    parser.add_argument("--output", required=False, help="输出xlsx路径")
    parser.add_argument("--token", required=False, help="直接使用 token (无需 CDP)")
    parser.add_argument("--company-id", required=False, help="直接使用 company ID (无需 CDP)")
    parser.add_argument("--company-name", required=False, default="凯旋创智测试集团", help="企业名称")
    parser.add_argument("--keep-group-inheritance", action="store_true", help="Sheet3分组按继承视觉输出")
    args = parser.parse_args()

    template = Path(args.template)
    if not template.exists():
        raise FileNotFoundError(f"模板不存在: {template}")

    browser = find_browser()
    if browser:
        print(f"✅ 检测到 {browser['name']} 浏览器 (端口 {browser['port']})")

    auth = get_auth_from_edge(
        token_override=args.token,
        company_id_override=args.company_id,
        company_name_override=args.company_name,
    )
    sources = fetch_sources(auth)
    
    print(f"📊 API中现有的一级科目: {sources['primary_subjects']}")

    out = Path(args.output) if args.output else template.parent / f"三表联动_客户模板_公司{auth.company_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    report_path = write_by_customer_template(
        template_path=template,
        out_path=out,
        auth=auth,
        sources=sources,
        inherit_group_visual=args.keep_group_inheritance,
    )

    print(f"✅ 生成完成: {out}")
    print(f"🧾 报告: {report_path}")


if __name__ == "__main__":
    main()
