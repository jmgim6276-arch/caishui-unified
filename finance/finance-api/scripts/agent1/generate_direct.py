#!/usr/bin/env python3
"""
Agent1: 基于客户模板生成三张表（01/02/03）
直接使用 token 和 companyId 运行，无需 CDP
"""

import argparse
import json
import random
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidationList
import requests
import websocket

BASE_URL = "https://cst.uf-tree.com"
# 直接设置 token 和 companyId
TOKEN = "2vERT8zu1Qem0vLtZ8Vre9a8pvY"
COMPANY_ID = 7792
COMPANY_NAME = "凯旋创智测试集团"

DOC_TYPES = ["报销单", "借款单", "批量付款单", "申请单"]


@dataclass
class Auth:
    token: str
    company_id: int
    company_name: str


# 直接返回 Auth 对象
def get_auth_from_edge() -> Auth:
    return Auth(
        token=TOKEN,
        company_id=COMPANY_ID,
        company_name=COMPANY_NAME,
    )


def api_get(auth: Auth, endpoint: str, params: Dict):
    return requests.get(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        params=params,
        timeout=15,
    ).json()


def api_post(auth: Auth, endpoint: str, payload: Dict):
    return requests.post(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        json=payload,
        timeout=15,
    ).json()


def fetch_sources(auth: Auth) -> Dict:
    # 1. 用户列表
    r = api_post(auth, "/api/member/department/queryCompany", {"companyId": auth.company_id})
    users = (r.get("result") or {}).get("users", []) or []
    
    # 2. 部门列表
    r = api_get(auth, "/api/member/department/queryDepartments", {"companyId": auth.company_id})
    departments = r.get("result", []) or []
    
    # 3. 角色列表
    r = api_get(auth, "/api/expense/role/queryRoleList", {"companyId": auth.company_id, "pageSize": 500})
    roles = r.get("result", {}).get("list", []) or []
    
    # 4. 费用科目（一级）
    r = api_get(auth, "/api/bill/feeTemplate/queryFeeTemplate", {"companyId": auth.company_id, "status": 1, "pageSize": 5000})
    all_fees = r.get("result", []) or []
    primary_subjects = [f for f in all_fees if f.get("parentId") == 0 or f.get("level") == 1]
    
    # 5. 工作流
    r = api_get(auth, "/api/flow/flowTemplate/queryList", {"companyId": auth.company_id, "status": 1})
    workflows = r.get("result", {}).get("list", []) or []
    
    return {
        "users": users,
        "departments": departments,
        "roles": roles,
        "primary_subjects": primary_subjects,
        "workflows": workflows,
    }


def unmerge_all(ws):
    """取消所有合并单元格"""
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))


def clear_data_validations(ws):
    """移除所有数据验证（下拉选项）"""
    ws.data_validations = DataValidationList()


def write_by_customer_template(template_path: Path, out_path: Path, auth: Auth, sources: Dict, inherit_group_visual: bool = False):
    wb = openpyxl.load_workbook(template_path)

    # 假设模板有这些 sheet
    expected_sheets = ["01-基础数据", "02-费用管控", "03-单据模板", "04-基础数据样本"]
    for s in expected_sheets:
        if s not in wb.sheetnames:
            print(f"警告: 模板缺少 sheet {s}")

    # 移除所有 sheet 的下拉选项（数据验证）
    for sheet_name in wb.sheetnames:
        clear_data_validations(wb[sheet_name])
    
    # 读取源数据
    users = sources.get("users", [])
    departments = sources.get("departments", [])
    roles = sources.get("roles", [])
    primary_subjects = sources.get("primary_subjects", [])
    
    # 生成 Sheet1 数据（基础数据）
    sheet1_rows = []
    for u in users[:50]:  # 限制数量
        sheet1_rows.append({
            "姓名": u.get("name", ""),
            "部门": u.get("departmentName", ""),
            "角色": u.get("roleName", ""),
        })
    
    # 生成 Sheet2 数据（费用管控）
    sheet2_rows = []
    for ps in primary_subjects[:20]:
        for doc_type in DOC_TYPES[:2]:
            sheet2_rows.append({
                "一级费用科目": ps.get("name", ""),
                "二级费用科目": "",
                "三级费用科目": "",
                "归属单据名称": doc_type,
                "单据适配人员": "",
                "备注": "",
            })
    
    # 生成 Sheet3 数据（单据模板）
    sheet3_rows = []
    for doc_type in DOC_TYPES:
        sheet3_rows.append({
            "单据分组（一级目录）": "通用",
            "单据大类（二级目录）": doc_type,
            "单据模板名称": doc_type,
            "可见范围类型": "全员",
            "可见范围对象": "",
            "备注": "",
        })
    
    # 写入数据
    if "01-基础数据" in wb.sheetnames:
        ws1 = wb["01-基础数据"]
        r = 2
        for row in sheet1_rows:
            ws1.cell(r, 1).value = row.get("姓名", "")
            ws1.cell(r, 2).value = row.get("部门", "")
            ws1.cell(r, 3).value = row.get("角色", "")
            r += 1
    
    if "02-费用管控" in wb.sheetnames:
        ws2 = wb["02-费用管控"]
        r = 2
        for row in sheet2_rows:
            ws2.cell(r, 1).value = row.get("一级费用科目", "")
            ws2.cell(r, 2).value = row.get("二级费用科目", "")
            ws2.cell(r, 3).value = row.get("三级费用科目", "")
            ws2.cell(r, 4).value = row.get("归属单据名称", "")
            ws2.cell(r, 5).value = row.get("单据适配人员", "")
            ws2.cell(r, 6).value = row.get("备注", "")
            r += 1
    
    if "03-单据模板" in wb.sheetnames:
        ws3 = wb["03-单据模板"]
        r = 2
        for row in sheet3_rows:
            ws3.cell(r, 1).value = row.get("单据分组（一级目录）", "")
            ws3.cell(r, 2).value = row.get("单据大类（二级目录）", "")
            ws3.cell(r, 3).value = row.get("单据模板名称", "")
            ws3.cell(r, 4).value = row.get("可见范围类型", "")
            ws3.cell(r, 5).value = row.get("可见范围对象", "")
            ws3.cell(r, 6).value = row.get("备注", "")
            r += 1

    # 保存前再次清除所有数据验证（确保标题行的下拉选项也被清除）
    # 同时移除所有自动筛选（只保留第二列的筛选稍后手动设置）
    for sheet_name in wb.sheetnames:
        clear_data_validations(wb[sheet_name])
        wb[sheet_name].auto_filter.ref = None

    # 为每个主要 sheet 设置只筛选第二列（B列：是否导入/是否执行/是否创建）
    if "01-基础数据" in wb.sheetnames:
        wb["01-基础数据"].auto_filter.ref = "B2"
    if "02-费用管控" in wb.sheetnames:
        wb["02-费用管控"].auto_filter.ref = "B2"
    if "03-单据模板" in wb.sheetnames:
        wb["03-单据模板"].auto_filter.ref = "B2"

    wb.save(out_path)
    
    # 生成报告
    report = {
        "companyId": auth.company_id,
        "companyName": auth.company_name,
        "output": str(out_path),
        "counts": {
            "users": len(users),
            "departments": len(departments),
            "roles": len(roles),
            "primary_subjects": len(primary_subjects),
            "sheet1_rows": len(sheet1_rows),
            "sheet2_rows": len(sheet2_rows),
            "sheet3_rows": len(sheet3_rows),
        },
    }
    
    report_path = out_path.with_suffix(".report.json")
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report_path


def main():
    parser = argparse.ArgumentParser(description="基于客户模板生成三表（Agent1）")
    parser.add_argument("--template", required=True, help="客户模板xlsx路径")
    parser.add_argument("--output", required=False, help="输出xlsx路径")
    parser.add_argument("--keep-group-inheritance", action="store_true", help="Sheet3分组按继承视觉输出")
    args = parser.parse_args()

    template = Path(args.template)
    if not template.exists():
        raise FileNotFoundError(f"模板不存在: {template}")

    auth = get_auth_from_edge()
    sources = fetch_sources(auth)

    out = Path(args.output) if args.output else template.parent / f"三表联动_客户模板_公司{auth.company_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    report_path = write_by_customer_template(
        template_path=template,
        out_path=out,
        auth=auth,
        sources=sources,
        inherit_group_visual=args.keep_group_inheritance,
    )

    print(f"✅ 生成完成: {out}")
    print(f"🧾 报告: {report_path}")


if __name__ == "__main__":
    main()
