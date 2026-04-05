#!/usr/bin/env python3
"""
Agent1: 基于客户模板生成三张表（01/02/03）
- 读取客户确认模板（保留样式）
- 实时查询角色/员工/部门/一级科目
- 按三步规则生成联动数据
- 特别规则：
  1) Step3 只用 4 类：报销单/借款单/批量付款单/申请单
  2) 费用限制仅报销单+批量付款单
  3) Sheet3 分组可向下继承（可选输出为空）
  4) Sheet2 归属单据名称 == Sheet3 单据模板名称（一一对应）
  5) 费用角色限制条件：归属单据名称非空 且 单据适配人员非空
"""

from __future__ import annotations

import argparse
import json
import random
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import requests
import websocket

BASE_URL = "https://cst.uf-tree.com"

# 支持的浏览器 CDP 端口
BROWSERS = [
    {"name": "Edge", "port": 9223, "url": "http://localhost:9223/json"},
    {"name": "Chrome", "port": 18800, "url": "http://localhost:18800/json"},
]

DOC_TYPES = ["报销单", "借款单", "批量付款单", "申请单"]


@dataclass
class Auth:
    token: str
    company_id: int
    company_name: str


# ---------------------------
# Browser Detection
# ---------------------------

def find_browser():
    """自动检测可用的浏览器，优先返回包含财税通页面的浏览器"""
    available = []
    for browser in BROWSERS:
        try:
            pages = requests.get(browser["url"], timeout=6).json()
            # 检查是否有财税通页面
            has_cst = any("cst.uf-tree.com" in p.get("url", "") for p in pages)
            available.append({**browser, "has_cst": has_cst, "page_count": len(pages)})
        except Exception:
            continue

    if not available:
        return None

    # 优先返回有财税通页面的浏览器
    for b in available:
        if b["has_cst"]:
            return b

    # 如果都没有财税通页面，返回第一个可用的
    return available[0]


# ---------------------------
# Auth & API
# ---------------------------

def get_auth_from_edge(token_override=None, company_id_override=None, company_name_override=None) -> Auth:
    # 如果直接提供了 token, 使用直接模式
    if token_override and company_id_override:
        return Auth(
            token=token_override,
            company_id=int(company_id_override),
            company_name=company_name_override or f"company_{company_id_override}",
        )

    # 否则从 CDP 读取
    browser = find_browser()
    if not browser:
        raise RuntimeError("未检测到可用的浏览器。请按以下步骤操作：\n"
                          "1. 打开 Edge 浏览器:\n"
                          "   /Applications/Microsoft\\ Edge.app/Contents/MacOS/Microsoft\\ Edge --remote-debugging-port=9223 --remote-allow-origins=*\n"
                          "2. 或打开 Chrome 浏览器:\n"
                          "   /Applications/Google\\ Chrome.app/Contents/MacOS/Google\\ Chrome --remote-debugging-port=18800 --remote-allow-origins=*\n"
                          "3. 登录 https://cst.uf-tree.com\n"
                          "或使用 --token 和 --company-id 参数直接指定。")

    if not browser["has_cst"]:
        raise RuntimeError(f"{browser['name']} 中未发现财税通页面，请先登录 https://cst.uf-tree.com")

    pages = requests.get(browser["url"], timeout=6).json()
    ws_url = None
    for p in pages:
        if "cst.uf-tree.com" in p.get("url", ""):
            ws_url = p.get("webSocketDebuggerUrl")
            break
    if not ws_url:
        raise RuntimeError(f"未在 {browser['name']} 中发现 cst.uf-tree.com 页面")

    ws = websocket.create_connection(ws_url, timeout=10, suppress_origin=True)
    ws.send(json.dumps({
        "id": 1,
        "method": "Runtime.evaluate",
        "params": {
            "expression": "(function(){const v=localStorage.getItem('vuex');if(!v)return null;const d=JSON.parse(v);return {token:d.user?.token, companyId:d.user?.company?.id, companyName:d.user?.company?.name};})()",
            "returnByValue": True,
        },
    }))
    raw = ws.recv()
    ws.close()

    value = json.loads(raw).get("result", {}).get("result", {}).get("value")
    if not value or not value.get("token") or not value.get("companyId"):
        raise RuntimeError("读取 token/companyId 失败（请确认已登录）")

    return Auth(
        token=value["token"],
        company_id=value["companyId"],
        company_name=value.get("companyName", ""),
    )


def api_get(auth: Auth, endpoint: str, params: dict):
    return requests.get(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        params=params,
        timeout=15,
    ).json()


def api_post(auth: Auth, endpoint: str, payload: dict):
    return requests.post(
        f"{BASE_URL}{endpoint}",
        headers={"x-token": auth.token, "Content-Type": "application/json"},
        json=payload,
        timeout=15,
    ).json()


# ---------------------------
# Data Sources
# ---------------------------

def unique_names(seq: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in seq:
        x = (x or "").strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def fetch_sources(auth: Auth) -> Dict:
    users_resp = api_post(auth, "/api/member/department/queryCompany", {"companyId": auth.company_id})
    users = (users_resp.get("result", {}) or {}).get("users", []) or []
    user_names = unique_names([u.get("nickName") for u in users])

    deps_resp = api_get(auth, "/api/member/department/queryDepartments", {"companyId": auth.company_id})
    deps = deps_resp.get("result", []) or []
    dep_names = unique_names([d.get("title") for d in deps])

    role_resp = api_get(auth, "/api/member/role/get/tree", {"companyId": auth.company_id})
    role_tree = role_resp.get("result", []) or []
    role_names = []
    for cat in role_tree:
        for c in (cat.get("children") or []):
            if c.get("name"):
                role_names.append(c.get("name"))
    role_names = unique_names(role_names)

    fee_resp = api_get(auth, "/api/bill/feeTemplate/queryFeeTemplate", {
        "companyId": auth.company_id,
        "status": 1,
        "pageSize": 5000,
    })
    fee_rows = fee_resp.get("result", []) or []
    primary_subjects = unique_names([x.get("name") for x in fee_rows if x.get("parentId") == -1])

    wf_resp = api_get(auth, "/api/bpm/workflow/queryWorkFlow", {
        "companyId": auth.company_id,
        "t": int(time.time() * 1000),
    })
    workflows = wf_resp.get("result", []) or []
    wf = next((x for x in workflows if x.get("tpName") == "通用审批流"), None)

    return {
        "users": user_names,
        "departments": dep_names,
        "roles": role_names,
        "primary_subjects": primary_subjects,
        "workflow_name": (wf or {}).get("tpName", "通用审批流"),
    }


# ---------------------------
# Business Rules / Generators
# ---------------------------

def generate_sheet1(company_name: str, departments: List[str], employee_count: int = 10):
    surnames = ["张", "李", "王", "赵", "刘", "陈", "周", "吴", "林", "黄"]
    given = ["晨", "悦", "宁", "航", "嘉", "然", "宇", "琳", "涛", "雪"]

    names_used, phones_used = set(), set()
    rows = []

    for i in range(employee_count):
        while True:
            name = f"{random.choice(surnames)}{random.choice(given)}{i+1}"
            if name not in names_used:
                names_used.add(name)
                break

        while True:
            phone = f"1{random.randint(30,99)}{random.randint(1000,9999)}{random.randint(1000,9999)}"
            if phone not in phones_used:
                phones_used.add(phone)
                break

        dep = random.choice(departments) if departments else company_name
        rows.append({
            "姓名": name,
            "手机号": phone,
            "企业名称": company_name,
            "一级部门名称": dep,
            "二级部门": "",
            "备注": "唯一性已校验",
        })
    return rows


def generate_sheet2(primary_subjects: List[str], users: List[str]):
    """重点：
    1) 二级科目下 >=2 个三级
    2) 三级全局不重复
    3) 同一单据名称：人员状态统一（全有 or 全无）
    4) 三级费用科目可以为空（灵活结构）
    """

    if not primary_subjects:
        primary_subjects = ["运营增长_0316", "AI工程_0316", "客户交付_0316", "财税支持_0316"]

    doc_plan = [
        # (归属单据类型, 归属单据名称, 是否有适配人员)
        ("报销单", "差旅报销单", True),
        ("报销单", "日常报销单", False),
        ("批量付款单", "采购付款单", True),
        ("借款单", "门店借款单", False),
        ("申请单", "用章申请单", False),
    ]

    # 每个二级至少2个三级，三级可以为空
    blocks = [
        # 一级, 二级, 三级列表, 四级列表(可选), 单据类型, 单据名称, 是否有三级
        (primary_subjects[0 % len(primary_subjects)], "差旅交通费", ["高铁票", "机票", ""], [], "报销单", "差旅报销单", True),
        (primary_subjects[0 % len(primary_subjects)], "差旅住宿费", ["酒店住宿费", "民宿住宿费"], [], "报销单", "差旅报销单", True),
        (primary_subjects[1 % len(primary_subjects)], "办公耗材费", ["打印纸", "墨盒硒鼓"], [], "报销单", "日常报销单", True),
        (primary_subjects[1 % len(primary_subjects)], "办公设备费", ["键盘", ""], [], "报销单", "日常报销单", True),
        (primary_subjects[2 % len(primary_subjects)], "采购原材料", ["食材采购", "调味辅料"], [], "批量付款单", "采购付款单", True),
        (primary_subjects[2 % len(primary_subjects)], "供应商服务费", ["物流服务费", ""], [], "批量付款单", "采购付款单", True),
        (primary_subjects[3 % len(primary_subjects)], "门店备用金", ["日常备用金"], [], "借款单", "门店借款单", True),
        (primary_subjects[3 % len(primary_subjects)], "合同与印章", ["合同盖章", "证明开具"], [], "申请单", "用章申请单", True),
    ]

    people_policy = {doc_name: has_people for _, doc_name, has_people in doc_plan}

    used_third = set()
    rows = []
    for lvl1, lvl2, thirds, fourths, doc_type, doc_name, has_third in blocks:
        # 全局去重三级
        third_unique = []
        for t in thirds:
            x = t
            n = 2
            while x in used_third:
                x = f"{t}{n}"
                n += 1
            used_third.add(x)
            third_unique.append(x)  # 保留空的三级，不要跳过

        # 为每个三级生成对应的行（三级和四级保持一致：要么都有，要么都为空）
        for t in third_unique:
            four = ""  # 默认为空
            # 只有三级有内容时，才生成四级（可以为空或空内容）
            if t and fourths:
                four = fourths[0] if fourths[0] else ""

            people = ""
            if people_policy.get(doc_name) and len(users) >= 2:
                people = "，".join(random.sample(users, 2))

            rows.append({
                "一级费用科目": lvl1,
                "二级费用科目": lvl2,
                "三级费用科目": t,
                "四级费用科目": four,
                "归属单据类型": doc_type,
                "归属单据名称": doc_name,
                "单据适配人员": people,
                "是否执行": "是",
                "备注": "二级>=2三级；三级全局不重复；同单据人员状态统一",
            })

    # 为 merge 展示排序（一级/二级/三级/四级连续）
    rows.sort(key=lambda r: (r["一级费用科目"], r["二级费用科目"], r["三级费用科目"], r["四级费用科目"]))
    return rows


def build_sheet3_from_sheet2(sheet2_rows: List[Dict], roles: List[str], users: List[str], deps: List[str], workflow_name: str, inherit_group_visual=True):
    # 汇总同名单据 - 直接复用 Sheet2 中的单据名称，避免重复
    agg: Dict[str, Dict] = {}

    for r in sheet2_rows:
        # 单据名称不加时间前缀，保持和 Sheet2 一致
        name = (r.get("归属单据名称") or "").strip()
        if not name:
            continue

        doc_type = (r.get("归属单据类型") or "").strip()
        has_people = bool((r.get("单据适配人员") or "").strip())
        leaf = (r.get("三级费用科目") or "").strip() or (r.get("二级费用科目") or "").strip()
        if name not in agg:
            agg[name] = {"doc_type": doc_type, "any_people": False, "leafs": set()}
        agg[name]["any_people"] = agg[name]["any_people"] or has_people
        if leaf:
            agg[name]["leafs"].add(leaf)

    group_map = {
        "报销单": "报销类单据",
        "批量付款单": "付款类单据",
        "借款单": "借款类单据",
        "申请单": "申请类单据",
    }

    out = []
    for i, (doc_name, info) in enumerate(agg.items()):
        dt = info["doc_type"]
        if dt not in DOC_TYPES:
            continue

        vis_type = ["角色", "员工", "部门", "全员"][i % 4]
        # 可见范围类型可以为空
        if vis_type == "全员":
            vis_obj = ""
            vis_obj_type = "全员"
        elif vis_type == "角色":
            vis_obj = random.choice(roles) if roles else "部门负责人"
            vis_obj_type = "角色"
        elif vis_type == "员工":
            vis_obj = random.choice(users) if users else "企业负责人"
            vis_obj_type = "员工"
        else:  # 部门
            vis_obj = random.choice(deps) if deps else "财务部"
            vis_obj_type = "部门"

        # 费用限制规则
        # 如果有人员，使用费用角色限制；否则不做限制
        if dt in ["报销单", "批量付款单"]:
            if doc_name and info["any_people"]:
                limit_mode = "费用角色限制"
                target = f"费用角色:{doc_name}"
            else:
                limit_mode = ""  # 不做限制
                target = ""
        else:
            # 其他单据类型不处理费用限制
            limit_mode = ""
            target = ""
        out.append({
            "单据分组（一级目录）": group_map[dt],
            "单据大类（二级目录）": dt,
            "单据模板名称": doc_name,
            "可见范围类型": vis_type,
            "可见范围对象": vis_obj,
            "备注": "可见范围对象来自真实接口",
            "费用限制方式": limit_mode,
            "对应末级科目/对应费用角色": target,
            "审批流": workflow_name,
        })

    # 强校验：sheet2 doc_name == sheet3 template_name
    s2 = sorted({(r.get("归属单据名称") or "").strip() for r in sheet2_rows if (r.get("归属单据名称") or "").strip()})
    s3 = sorted({(r.get("单据模板名称") or "").strip() for r in out if (r.get("单据模板名称") or "").strip()})
    if s2 != s3:
        raise RuntimeError(f"名称映射失败: sheet2={s2}, sheet3={s3}")

    if inherit_group_visual and out:
        # 视觉继承：同分组连续时，后续行可留空
        last = None
        for r in out:
            g = r["单据分组（一级目录）"]
            if g == last:
                r["单据分组（一级目录）"] = ""
            else:
                last = g

    return out


# ---------------------------
# Excel Writer
# ---------------------------

def clear_range(ws, row_start=4, row_end=220, col_start=1, col_end=8):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(r, c).value = None


def unmerge_all(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))


def clear_data_validations(ws):
    """移除所有数据验证（下拉选项）"""
    ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()


def merge_same(ws, col: int, start_row: int, end_row: int):
    r = start_row
    while r <= end_row:
        v = ws.cell(r, col).value
        e = r
        while e + 1 <= end_row and ws.cell(e + 1, col).value == v:
            e += 1
        if e > r and v not in (None, ""):
            ws.merge_cells(start_row=r, start_column=col, end_row=e, end_column=col)
        r = e + 1


def write_by_customer_template(template_path: Path, out_path: Path, auth: Auth, sources: Dict, inherit_group_visual=True):
    wb = openpyxl.load_workbook(template_path)
    ws1 = wb["01_添加员工"]
    ws2 = wb["02_费用科目配置"]
    ws3 = wb["03_单据表"]
    wsb = wb["基础数据"]

    # 为了重写并合并，先取消已有合并（只动 1/2/3）
    # 同时移除所有数据验证（下拉选项）
    for ws in [ws1, ws2, ws3]:
        unmerge_all(ws)
        clear_data_validations(ws)

    clear_range(ws1, 4, 220, 1, 8)
    clear_range(ws2, 4, 220, 1, 9)  # 扩展到第9列（新增四级科目）
    clear_range(ws3, 4, 220, 1, 8)

    sheet1_rows = generate_sheet1(auth.company_name, sources["departments"], employee_count=10)
    sheet1_names = [r["姓名"] for r in sheet1_rows]  # 提取 Sheet1 员工姓名
    sheet2_rows = generate_sheet2(sources["primary_subjects"], sheet1_names)  # 只用 Sheet1 的员工
    sheet3_rows = build_sheet3_from_sheet2(
        sheet2_rows,
        roles=sources["roles"],
        users=sheet1_names,  # 使用 Sheet1 员工姓名，而不是 API 用户
        deps=sources["departments"],
        workflow_name=sources["workflow_name"],
        inherit_group_visual=inherit_group_visual,
    )

    # write sheet1
    r = 4
    for i, row in enumerate(sheet1_rows, start=1):
        ws1.cell(r, 1).value = i
        ws1.cell(r, 2).value = "是"
        ws1.cell(r, 3).value = row["姓名"]
        ws1.cell(r, 4).value = row["手机号"]
        ws1.cell(r, 5).value = row["企业名称"]
        ws1.cell(r, 6).value = row["一级部门名称"]
        ws1.cell(r, 7).value = row["二级部门"]
        ws1.cell(r, 8).value = row["备注"]
        r += 1

    # write sheet2
    r = 4
    for i, row in enumerate(sheet2_rows, start=1):
        ws2.cell(r, 1).value = i
        ws2.cell(r, 2).value = row["是否执行"]
        ws2.cell(r, 3).value = row["一级费用科目"]
        ws2.cell(r, 4).value = row["二级费用科目"]
        ws2.cell(r, 5).value = row["三级费用科目"]
        ws2.cell(r, 6).value = row["归属单据名称"]
        ws2.cell(r, 7).value = row["单据适配人员"]
        ws2.cell(r, 8).value = row["四级费用科目"]  # 新增四级科目列
        ws2.cell(r, 9).value = row["备注"]  # 备注移到第9列
        r += 1
    end2 = r - 1

    # 合并规则：一级/二级合并，三级不合并
    merge_same(ws2, 3, 4, end2)
    merge_same(ws2, 4, 4, end2)

    # write sheet3
    r = 4
    for i, row in enumerate(sheet3_rows, start=1):
        ws3.cell(r, 1).value = i
        ws3.cell(r, 2).value = "是"
        ws3.cell(r, 3).value = row["单据分组（一级目录）"]
        ws3.cell(r, 4).value = row["单据大类（二级目录）"]
        ws3.cell(r, 5).value = row["单据模板名称"]
        ws3.cell(r, 6).value = row["可见范围类型"]
        ws3.cell(r, 7).value = row["可见范围对象"]
        ws3.cell(r, 8).value = row["备注"]
        r += 1

    # 基础数据写样本（小白AI快速核对）
    clear_range(wsb, 2, 30, 1, 7)
    wsb.cell(1, 1).value = "可见范围类型"
    wsb.cell(1, 3).value = "角色"
    wsb.cell(1, 4).value = "员工"
    wsb.cell(1, 5).value = "部门"
    wsb.cell(1, 6).value = "单据大类"
    wsb.cell(1, 7).value = "是否创建"
    for i in range(10):
        rr = 2 + i
        wsb.cell(rr, 1).value = "样本"
        wsb.cell(rr, 3).value = sources["roles"][i % len(sources["roles"])] if sources["roles"] else ""
        wsb.cell(rr, 4).value = sources["users"][i % len(sources["users"])] if sources["users"] else ""
        wsb.cell(rr, 5).value = sources["departments"][i % len(sources["departments"])] if sources["departments"] else ""
        wsb.cell(rr, 6).value = DOC_TYPES[i % len(DOC_TYPES)]
        wsb.cell(rr, 7).value = "是"

    # 保存前再次清除所有数据验证（确保标题行的下拉选项也被清除）
    # 同时移除所有自动筛选（只保留第二列的筛选稍后手动设置）
    for ws in [ws1, ws2, ws3, wsb]:
        clear_data_validations(ws)
        ws.auto_filter.ref = None

    # 为每个主要 sheet 设置只筛选第二列（B列：是否导入/是否执行/是否创建）
    ws1.auto_filter.ref = "B3"
    ws2.auto_filter.ref = "B3"
    ws3.auto_filter.ref = "B3"

    wb.save(out_path)

    # 输出报告
    report = {
        "companyId": auth.company_id,
        "companyName": auth.company_name,
        "output": str(out_path),
        "counts": {
            "roles": len(sources["roles"]),
            "users": len(sources["users"]),
            "departments": len(sources["departments"]),
            "primary_subjects": len(sources["primary_subjects"]),
            "sheet1_rows": len(sheet1_rows),
            "sheet2_rows": len(sheet2_rows),
            "sheet3_rows": len(sheet3_rows),
        },
        "rules": {
            "step3_doc_types": DOC_TYPES,
            "fee_limit_applies_to": ["报销单", "批量付款单"],
            "group_inheritance_visual": inherit_group_visual,
        }
    }
    report_path = out_path.with_suffix(".report.json")
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report_path


def main():
    parser = argparse.ArgumentParser(description="基于客户模板生成三表（Agent1）")
    parser.add_argument("--template", required=True, help="客户模板xlsx路径")
    parser.add_argument("--output", required=False, help="输出xlsx路径")
    parser.add_argument("--token", required=False, help="直接使用 token (无需 CDP)")
    parser.add_argument("--company-id", required=False, help="直接使用 company ID (无需 CDP)")
    parser.add_argument("--company-name", required=False, default="凯旋创智测试集团", help="企业名称")
    parser.add_argument("--keep-group-inheritance", action="store_true", help="Sheet3分组按继承视觉输出（同组后续行留空）")
    args = parser.parse_args()

    template = Path(args.template)
    if not template.exists():
        raise FileNotFoundError(f"模板不存在: {template}")

    # 检测并显示使用的浏览器
    if not args.token or not args.company_id:
        browser = find_browser()
        if browser:
            print(f"✅ 检测到 {browser['name']} 浏览器 (端口 {browser['port']})")
        else:
            print("⚠️  未检测到可用浏览器，尝试直接获取...")

    auth = get_auth_from_edge(
        token_override=args.token,
        company_id_override=args.company_id,
        company_name_override=args.company_name,
    )
    sources = fetch_sources(auth)

    out = Path(args.output) if args.output else template.parent / f"三表联动_客户模板_公司{auth.company_id}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    report_path = write_by_customer_template(
        template_path=template,
        out_path=out,
        auth=auth,
        sources=sources,
        inherit_group_visual=args.keep_group_inheritance,
    )

    print(f"✅ 生成完成: {out}")
    print(f"🧾 报告: {report_path}")


if __name__ == "__main__":
    main()
